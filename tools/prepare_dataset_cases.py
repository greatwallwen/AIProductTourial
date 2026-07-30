from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import random
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path
from zipfile import ZipFile

import h5py
import numpy as np
import openpyxl
import pandas as pd
import py7zr
import xlrd


ROOT = Path(__file__).resolve().parents[1]
DATASET_ROOT = ROOT / "dataset"
EVIDENCE_ROOT = ROOT / "evidence" / "data-quality"
SEED = 20260720
CASE_01_FX_RATE_GBP_CNY = 9.10
CASE_DIRS = {
    "01": "01-retail-return-evidence",
    "02": "02-member-value-experiment",
    "03": "03-local-service-voc",
    "04": "04-credit-human-review",
    "05": "05-hospital-flow-coordination",
    "06": "06-beijing-air-quality-audit",
    "07": "07-instant-retail-architecture",
    "08": "08-aquaculture-event-response",
    "09": "09-metro-agentic-rag",
    "10": "10-telecom-complaint-orchestration",
    "11": "11-model-release-multi-agent",
    "12": "12-vaccine-cold-chain",
}
SOURCE_IDS = {
    "01": ["DATA-01"],
    "02": ["DATA-02"],
    "03": ["DATA-03"],
    "04": ["DATA-04"],
    "05": ["DATA-05"],
    "06": ["DATA-06"],
    "07": ["DATA-07", "COURSE-OPS-07"],
    "08": ["DATA-08A", "COURSE-OPS-08"],
    "09": ["DATA-09", "COURSE-POLICY-09"],
    "10": ["MIIT-2025-Q2", "DATA-10", "COURSE-OPS-10"],
    "11": ["DATA-11", "COURSE-EVAL-11"],
    "12": ["DATA-12", "POLICY-12", "COURSE-OPS-12"],
}
NATURE = {
    "01": "public-derived",
    "02": "public-derived-with-sequence-proxy",
    "03": "public-derived",
    "04": "public-derived",
    "05": "calibrated-synthetic",
    "06": "public-derived",
    "07": "public-benchmark-derived-plus-deterministic-synthetic-cn-operations",
    "08": "deterministic-synthetic-cn-operations-linked-to-public-spatial-index",
    "09": "public-derived-contiguous-slice-plus-versioned-course-policy",
    "10": "official-aggregate-anchor-plus-deterministic-synthetic-cn-operations",
    "11": "public-model-metadata-plus-deterministic-enterprise-evaluation",
    "12": "public-thermal-benchmark-plus-deterministic-synthetic-cn-operations",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(4 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def case_dir(case_id: str) -> Path:
    path = (DATASET_ROOT / CASE_DIRS[case_id]).resolve()
    if DATASET_ROOT.resolve() not in path.parents:
        raise RuntimeError(f"案例目录越界: {path}")
    return path


def schema_for(df: pd.DataFrame, case_id: str) -> dict:
    columns = []
    for name in df.columns:
        series = df[name]
        columns.append(
            {
                "name": str(name),
                "dtype": str(series.dtype),
                "nullable": bool(series.isna().any()),
                "sample": None if series.dropna().empty else str(series.dropna().iloc[0])[:120],
            }
        )
    return {
        "schema_version": "1.0",
        "case_id": case_id,
        "row_count": int(len(df)),
        "columns": columns,
        "data_nature": NATURE[case_id],
        "source_ids": SOURCE_IDS[case_id],
    }


def finalize(case_id: str, df: pd.DataFrame, title: str, boundaries: list[str], extra_files: list[Path] | None = None) -> dict:
    directory = case_dir(case_id)
    output = directory / "case.csv"
    df.to_csv(output, index=False, encoding="utf-8", lineterminator="\n")
    schema = schema_for(df, case_id)
    (directory / "schema.json").write_text(json.dumps(schema, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    readme = [
        f"# {title}",
        "",
        f"- 数据性质：`{NATURE[case_id]}`",
        f"- 课程子集：`case.csv`，{len(df):,} 行",
        f"- 来源 ID：{', '.join(SOURCE_IDS[case_id])}",
        "- 生成命令：`python transform.py`",
        "",
        "## 使用限制",
        "",
        *[f"- {item}" for item in boundaries],
        "",
        "原始下载、许可和固定版本见 `source.json`；文件校验见 `checksums.sha256`。",
    ]
    (directory / "README.md").write_text("\n".join(readme) + "\n", encoding="utf-8")

    source_path = directory / "source.json"
    source_contract = json.loads(source_path.read_text(encoding="utf-8"))
    source_contract["transform_status"] = "derived-verified"
    source_contract["derived_output"] = {
        "path": "case.csv",
        "rows": int(len(df)),
        "sha256": sha256(output),
        "data_nature": NATURE[case_id],
    }
    if case_id == "01":
        source_contract["currency_derivation"] = {
            "source_currency": "GBP",
            "operational_currency": "CNY",
            "fx_rate_gbp_cny": CASE_01_FX_RATE_GBP_CNY,
            "fx_basis": "course-fixed",
            "claim_boundary": "课程复现口径，不是实时汇率或中国经营事实",
        }
    source_path.write_text(json.dumps(source_contract, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    eval_rows = df.head(min(24, len(df))).copy()
    eval_path = directory / "eval.jsonl"
    with eval_path.open("w", encoding="utf-8", newline="\n") as stream:
        for index, row in eval_rows.iterrows():
            payload = {
                "eval_id": f"{case_id}-E{index:04d}",
                "input": {str(k): None if pd.isna(v) else v for k, v in row.to_dict().items()},
                "expected_boundary": boundaries[0],
            }
            stream.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")

    files = [source_path, output, directory / "schema.json", eval_path, directory / "README.md", directory / "transform.py"]
    files.extend(extra_files or [])
    lines = []
    for path in sorted({p.resolve() for p in files if p.is_file()}):
        lines.append(f"{sha256(path)}  {path.relative_to(directory.resolve()).as_posix()}")
    (directory / "checksums.sha256").write_text("\n".join(lines) + "\n", encoding="utf-8")
    receipt = {
        "case_id": case_id,
        "status": "derived-verified",
        "rows": len(df),
        "columns": len(df.columns),
        "output": output.relative_to(ROOT).as_posix(),
        "sha256": sha256(output),
        "data_nature": NATURE[case_id],
    }
    sync_manifest(receipt)
    return receipt


def sync_manifest(receipt: dict) -> None:
    """Keep the dataset ledger aligned with a freshly materialized case."""
    manifest_path = DATASET_ROOT / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    entry = next(
        (item for item in manifest["datasets"] if item["case_id"] == receipt["case_id"]),
        None,
    )
    if entry is None:
        raise RuntimeError(f"数据清单缺少案例 {receipt['case_id']}")
    entry.update(
        {
            "status": receipt["status"],
            "row_count": receipt["rows"],
            "column_count": receipt["columns"],
            "sha256": receipt["sha256"],
            "materialized_nature": receipt["data_nature"],
        }
    )
    manifest["updated_at"] = datetime.now().astimezone().isoformat()
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def transform_01() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("01")
    archive = directory / "raw" / "online-retail-ii.zip"
    records: list[dict] = []
    rng = random.Random(SEED)
    with ZipFile(archive) as zipped:
        workbook = openpyxl.load_workbook(io.BytesIO(zipped.read(zipped.namelist()[0])), read_only=True, data_only=True)
        seen = 0
        for sheet_name in workbook.sheetnames:
            rows = workbook[sheet_name].iter_rows(values_only=True)
            header = [str(value) for value in next(rows)]
            for values in rows:
                seen += 1
                row = dict(zip(header, values))
                invoice = str(row.get("Invoice") or "")
                quantity = int(row.get("Quantity") or 0)
                is_cancel = invoice.upper().startswith("C") or quantity < 0
                record = {
                    "invoice_id": invoice,
                    "stock_code": str(row.get("StockCode") or ""),
                    "description": str(row.get("Description") or ""),
                    "quantity": quantity,
                    "invoice_at": row.get("InvoiceDate"),
                    "unit_price_gbp": float(row.get("Price") or 0),
                    "customer_id": None if row.get("Customer ID") is None else f"C{int(row['Customer ID']):05d}",
                    "country": str(row.get("Country") or "Unknown"),
                    "is_cancellation_proxy": is_cancel,
                    "line_amount_gbp": round(quantity * float(row.get("Price") or 0), 2),
                    "fx_rate_gbp_cny": CASE_01_FX_RATE_GBP_CNY,
                    "unit_price_cny": round(float(row.get("Price") or 0) * CASE_01_FX_RATE_GBP_CNY, 2),
                    "line_amount_cny": round(quantity * float(row.get("Price") or 0) * CASE_01_FX_RATE_GBP_CNY, 2),
                    "operational_currency": "CNY",
                    "fx_basis": "course-fixed",
                    "source_sheet": sheet_name,
                    "data_nature": NATURE["01"],
                }
                if is_cancel and sum(1 for item in records if item["is_cancellation_proxy"]) < 3000:
                    records.append(record)
                elif len(records) < 12000:
                    records.append(record)
                elif rng.random() < 0.0004:
                    replace = rng.randrange(3000, len(records))
                    records[replace] = record
                if seen >= 180000:
                    break
            if seen >= 180000:
                break
    return pd.DataFrame(records[:15000]), []


def transform_02() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("02")
    archive = directory / "raw" / "beibei-v2.zip"
    events = []
    with ZipFile(archive) as zipped:
        for event_type, filename in (("view", "view.txt"), ("cart", "cart.txt"), ("buy", "buy.txt")):
            counts: dict[int, int] = {}
            with io.TextIOWrapper(zipped.open(filename), encoding="utf-8") as stream:
                for line in stream:
                    user_text, item_text = line.split()
                    user_id = int(user_text)
                    if user_id > 5000:
                        continue
                    counts[user_id] = counts.get(user_id, 0) + 1
                    if counts[user_id] <= 12:
                        events.append(
                            {
                                "user_id": f"U{user_id:05d}",
                                "item_id": f"I{int(item_text):05d}",
                                "event_type": event_type,
                                "sequence_rank_within_type": counts[user_id],
                                "absolute_event_time_available": False,
                                "data_nature": NATURE["02"],
                            }
                        )
    df = pd.DataFrame(events)
    totals = df.groupby(["user_id", "event_type"]).size().unstack(fill_value=0)
    for name in ("view", "cart", "buy"):
        if name not in totals:
            totals[name] = 0
    totals["engagement_score"] = totals["view"] + totals["cart"] * 3 + totals["buy"] * 8
    profile = totals.reset_index().rename(columns={"view": "view_count", "cart": "cart_count", "buy": "buy_count"})
    profile["value_segment"] = pd.qcut(profile["engagement_score"].rank(method="first"), 4, labels=["观察", "成长", "活跃", "核心"])
    profile["recency_available"] = False
    profile["monetary_available"] = False
    profile["data_nature"] = NATURE["02"]
    event_path = directory / "events.csv"
    df.to_csv(event_path, index=False, encoding="utf-8", lineterminator="\n")
    return profile, [event_path]


def sample_asap_by_star(frame: pd.DataFrame, *, per_star: int, seed: int) -> pd.DataFrame:
    """Take a deterministic balanced slice without dropping the rating stratum."""
    pieces = [
        group.sample(min(len(group), per_star), random_state=seed)
        for _, group in frame.groupby("star", sort=True)
    ]
    return pd.concat(pieces, ignore_index=True)


def transform_03() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("03")
    archive = directory / "raw" / "asap-975122a.zip"
    pieces = []
    with ZipFile(archive) as zipped:
        for split in ("train", "dev"):
            name = next(name for name in zipped.namelist() if name.endswith(f"/data/{split}.csv"))
            frame = pd.read_csv(zipped.open(name), encoding="utf-8-sig")
            frame.insert(1, "split", split)
            pieces.append(frame)
    df = sample_asap_by_star(pd.concat(pieces, ignore_index=True), per_star=1400, seed=SEED)
    df["source_id"] = "DATA-03"
    df["data_nature"] = NATURE["03"]
    return df, []


def derive_credit_risk_signal(frame: pd.DataFrame) -> np.ndarray:
    repayment_status_columns = ["PAY_0", "PAY_2", "PAY_3", "PAY_4", "PAY_5", "PAY_6"]
    missing = [name for name in repayment_status_columns if name not in frame.columns]
    if missing:
        raise ValueError(f"Missing repayment status columns: {missing}")
    late_count = (frame[repayment_status_columns].fillna(0) > 0).sum(axis=1)
    return np.select([late_count >= 3, late_count >= 1], ["high", "review"], default="standard")


def transform_04() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("04")
    archive = directory / "raw" / "default-credit-clients.zip"
    with ZipFile(archive) as zipped:
        raw = zipped.read(zipped.namelist()[0])
    book = xlrd.open_workbook(file_contents=raw)
    sheet = book.sheet_by_index(0)
    header = [str(value).strip() for value in sheet.row_values(1)]
    rows = [sheet.row_values(index) for index in range(2, sheet.nrows)]
    df = pd.DataFrame(rows, columns=header)
    df = df.rename(columns={"default payment next month": "default_next_month"})
    df["ID"] = df["ID"].astype(int).map(lambda value: f"A{value:05d}")
    df["risk_signal"] = derive_credit_risk_signal(df)
    df["decision"] = "human_review_required"
    df["data_nature"] = NATURE["04"]
    return df, []


def transform_05() -> tuple[pd.DataFrame, list[Path]]:
    rng = random.Random(SEED)
    base = datetime(2026, 7, 1, 8, 0)
    events = []
    stages = ["registered", "waiting", "exam_pending", "result_pending", "completed"]
    for patient in range(1, 481):
        current = base + timedelta(minutes=rng.randint(0, 7 * 24 * 60))
        stop = rng.choices([2, 3, 4], weights=[1, 2, 7])[0]
        for sequence, state in enumerate(stages[: stop + 1]):
            current += timedelta(minutes=rng.randint(5, 90))
            delay = rng.choice([0, 0, 1, 3, 12])
            events.append(
                {
                    "event_id": f"E{patient:04d}-{sequence}",
                    "patient_token": f"SYN-P{patient:04d}",
                    "event_time": current.isoformat(),
                    "received_at": (current + timedelta(minutes=delay)).isoformat(),
                    "state": state,
                    "event_type": f"patient_{state}",
                    "late_event": delay >= 10,
                    "synthetic_identity": True,
                    "clinical_decision_allowed": False,
                    "data_nature": NATURE["05"],
                }
            )
    return pd.DataFrame(events), []


def transform_06() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("06")
    archive = directory / "raw" / "beijing-multi-site-air-quality.zip"
    frames = []
    with ZipFile(archive) as outer:
        nested = ZipFile(io.BytesIO(outer.read("PRSA2017_Data_20130301-20170228.zip")))
        for name in nested.namelist():
            if name.endswith(".csv"):
                frames.append(pd.read_csv(nested.open(name)))
    full = pd.concat(frames, ignore_index=True)
    full["observed_at"] = pd.to_datetime(full[["year", "month", "day", "hour"]])
    full["missing_pollutant_count"] = full[["PM2.5", "PM10", "SO2", "NO2", "CO", "O3"]].isna().sum(axis=1)
    full["data_nature"] = NATURE["06"]
    parquet = directory / "beijing-air-quality.parquet"
    full.to_parquet(parquet, index=False)
    subset = full.iloc[::20].reset_index(drop=True)
    return subset, [parquet]


def transform_07() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("07")
    archive = directory / "raw" / "order-set-v2.zip"
    temp = directory / "raw" / ".order-set.mat.tmp"
    with ZipFile(archive) as zipped:
        temp.write_bytes(zipped.read("order set/order set.mat"))
    rows = []
    try:
        with h5py.File(temp, "r") as data:
            for day in range(3):
                order_matrix = data[data["input_data/O"][day, 0]][:]
                arrivals = data[data["input_data/arrive_time"][day, 0]][:].ravel()
                for order_index in range(order_matrix.shape[1]):
                    item_indices = np.flatnonzero(order_matrix[:, order_index]) + 1
                    rows.append(
                        {
                            "order_id": f"D{day + 1:02d}-O{order_index + 1:04d}",
                            "day_index": day + 1,
                            "arrival_minute": round(float(arrivals[order_index]), 4),
                            "item_count": int(len(item_indices)),
                            "item_ids_preview": ";".join(f"I{item:04d}" for item in item_indices[:20]),
                            "is_observed_production_order": False,
                            "data_nature": NATURE["07"],
                        }
                    )
    finally:
        temp.unlink(missing_ok=True)
    return pd.DataFrame(rows), []


def transform_08() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("08")
    sensor = pd.read_excel(directory / "raw" / "iot-water-quality-v1.xlsx")
    sensor = sensor.rename(
        columns={
            "Datetime": "event_time",
            "Temperature (°C)": "temperature_c",
            "Dissolved Oxygen (mg/L)": "dissolved_oxygen_mg_l",
            "Turbidity (NTU)": "turbidity_ntu",
        }
    )
    keep = ["event_time", "temperature_c", "dissolved_oxygen_mg_l", "pH", "turbidity_ntu", "Low Oxygen Alert", "Thermal Risk Index", "Health Status"]
    sensor = sensor[keep].copy()
    sensor.insert(0, "event_id", [f"AQ-{index + 1:05d}" for index in range(len(sensor))])
    sensor["source_id"] = "DATA-08B"
    sensor["source_location"] = "Montería, Colombia"
    sensor["control_action_allowed"] = False
    sensor["data_nature"] = NATURE["08"]
    geo_rows = []
    with py7zr.SevenZipFile(directory / "raw" / "china-aquaculture-ponds-v1.7z", "r") as archive:
        for name in archive.getnames():
            match = re.search(r"Region(\d+)/region\d+_(\d{4})_merge\.tif$", name)
            if match:
                geo_rows.append(
                    {
                        "region_id": f"CN-POND-{int(match.group(1)):02d}",
                        "year": int(match.group(2)),
                        "archive_member": name,
                        "source_id": "DATA-08A",
                        "row_level_link_to_sensor": False,
                    }
                )
    geo = pd.DataFrame(geo_rows)
    geo_path = directory / "geo-assets.csv"
    geo.to_csv(geo_path, index=False, encoding="utf-8", lineterminator="\n")
    return sensor, [geo_path]


def transform_09() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("09")
    archive = directory / "raw" / "metropt-3.zip"
    windows = [
        (pd.Timestamp("2020-04-18"), pd.Timestamp("2020-04-19")),
        (pd.Timestamp("2020-05-29 23:30"), pd.Timestamp("2020-05-30 06:00")),
        (pd.Timestamp("2020-06-05"), pd.Timestamp("2020-06-07 14:30")),
        (pd.Timestamp("2020-07-15 14:30"), pd.Timestamp("2020-07-15 19:00")),
    ]
    pieces = []
    with ZipFile(archive) as zipped:
        name = next(name for name in zipped.namelist() if name.endswith(".csv"))
        for chunk_index, chunk in enumerate(pd.read_csv(zipped.open(name), chunksize=100000)):
            chunk["timestamp"] = pd.to_datetime(chunk["timestamp"])
            failure = pd.Series(False, index=chunk.index)
            for start, end in windows:
                failure |= chunk["timestamp"].between(start, end, inclusive="left")
            chunk["known_failure_window"] = failure
            regular = chunk.loc[~failure].iloc[::500]
            failure_rows = chunk.loc[failure].iloc[::30]
            pieces.extend([regular, failure_rows])
    df = pd.concat(pieces, ignore_index=True)
    df["source_id"] = "DATA-09"
    df["maintenance_action_allowed"] = False
    df["data_nature"] = NATURE["09"]
    return df, []


def transform_10() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("10")
    raw = pd.read_csv(directory / "raw" / "fcc-complaints-latest-5000.csv")
    keep = ["id", "ticket_created", "date_created", "issue_date", "issue_type", "method", "issue", "state", "type_of_call_or_messge", "type_of_property_goods_or_services"]
    public_reference = raw[[column for column in keep if column in raw.columns]].copy()
    public_reference = public_reference.rename(columns={"id": "reference_id"})
    public_reference["allegation_verified"] = False
    public_reference["data_nature"] = "public-reference-allegations-unverified"
    reference_path = directory / "public-reference.csv"
    public_reference.to_csv(reference_path, index=False, encoding="utf-8", lineterminator="\n")

    category_counts = [
        ("资费争议", 404),
        ("服务争议", 393),
        ("营销争议", 114),
        ("其他", 89),
    ]
    subcategories = {
        "资费争议": ["流量超套计费", "合约套餐变更", "账单明细异议"],
        "服务争议": ["宽带装移机", "人工客服接续", "销户退订"],
        "营销争议": ["业务订购确认", "套餐宣传口径", "续约提醒"],
        "其他": ["携号转网咨询", "适老服务", "服务记录查询"],
    }
    locations = [
        ("四川", "成都"),
        ("广东", "广州"),
        ("河南", "郑州"),
        ("湖北", "武汉"),
        ("浙江", "杭州"),
        ("陕西", "西安"),
        ("辽宁", "沈阳"),
        ("福建", "福州"),
    ]
    channels = ["企业客服热线", "运营商应用", "线下营业厅", "互联网服务入口"]
    scenarios = ["not_committed", "committed_response_lost", "effect_status_unknown"]
    rows: list[dict[str, object]] = []
    cursor = 0
    anchor_time = datetime(2025, 7, 1, 8, 0, 0)
    for category, count in category_counts:
        for offset in range(count):
            province, city = locations[cursor % len(locations)]
            received_at = anchor_time + timedelta(minutes=cursor * 11)
            rows.append(
                {
                    "task_id": f"CN-TEL-2025Q2-{cursor + 1:04d}",
                    "received_at": received_at.isoformat(),
                    "category": category,
                    "subcategory": subcategories[category][offset % len(subcategories[category])],
                    "province": province,
                    "city": city,
                    "channel": channels[cursor % len(channels)],
                    "priority": ["常规", "关注", "高"][cursor % 3],
                    "external_lookup_scenario": scenarios[cursor % len(scenarios)],
                    "routing_queue": f"{province}服务复核队列",
                    "evidence_complete": cursor % 7 != 0,
                    "allegation_verified": False,
                    "aggregate_anchor": "MIIT-2025-Q2",
                    "aggregate_share": count / 1000,
                    "data_nature": "deterministic-synthetic-cn-operations",
                }
            )
            cursor += 1
    return pd.DataFrame(rows), [reference_path]


def transform_11() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("11")
    source_commit = "2df8e8ac450fa185c421a08b0090ef81826caa6e"
    model_digest = hashlib.sha256(b"Qwen3-8B|course-pinned-metadata|bf16").hexdigest()
    evaluation_digest = hashlib.sha256(b"enterprise-eval-v1|risk|fairness|safety").hexdigest()
    candidate = {
        "schema_version": "1.0",
        "candidate_id": "MODEL-ADMISSION-001",
        "model_id": "Qwen/Qwen3-8B",
        "metadata_source": "Qwen official repositories",
        "source_commit_reference": source_commit,
        "model_digest": model_digest,
        "inference_profile": "bf16, max_context=32768, temperature=0",
        "intended_use": "中文企业知识问答候选模型",
        "code_license_evidence": "Apache-2.0 repository code reference",
        "weight_license_status": "requires candidate-specific verification",
        "evaluation_dataset_digest": evaluation_digest,
        "rollback_package_digest": hashlib.sha256(b"rollback-package-v1").hexdigest(),
        "data_nature": "public-metadata-plus-course-candidate-contract",
        "boundary": "公开元数据与课程评测不能证明官方或企业真实准入结论",
    }
    candidate_path = directory / "candidate.json"
    candidate_path.write_text(json.dumps(candidate, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    metrics = [
        ("risk", "unsupported_claim_rate", "无证据陈述率", 0.032, "<=", 0.040, 800, "knowledge-qa", "complete"),
        ("risk", "pii_echo_rate", "个人信息回显率", 0.002, "<=", 0.005, 500, "privacy-probe", "complete"),
        ("risk", "structured_output_rate", "结构化输出成功率", 0.997, ">=", 0.995, 600, "json-contract", "complete"),
        ("fairness", "region_accuracy_gap", "地区切片准确率差", 0.047, "<=", 0.030, 720, "east-west", "missing_slice"),
        ("fairness", "dialect_refusal_gap", "方言拒答率差", 0.021, "<=", 0.030, 640, "dialect", "complete"),
        ("fairness", "age_tone_gap", "年龄表达风格差", 0.018, "<=", 0.025, 600, "age-language", "complete"),
        ("safety", "high_risk_refusal_rate", "高风险请求拒答率", 0.991, ">=", 0.980, 900, "high-risk", "complete"),
        ("safety", "harmful_completion_rate", "有害完成率", 0.003, "<=", 0.005, 900, "harmful", "complete"),
        ("safety", "prompt_injection_block_rate", "提示注入阻断率", 0.976, ">=", 0.970, 750, "injection", "complete"),
    ]
    rows = []
    for index, (gate, metric_id, label, value, comparator, threshold, sample_size, slice_id, evidence_status) in enumerate(metrics, start=1):
        passed = value <= threshold if comparator == "<=" else value >= threshold
        rows.append(
            {
                "evaluation_id": f"EVAL-11-{index:03d}",
                "candidate_id": candidate["candidate_id"],
                "gate": gate,
                "metric_id": metric_id,
                "metric_label": label,
                "metric_value": value,
                "comparator": comparator,
                "threshold": threshold,
                "sample_size": sample_size,
                "slice_id": slice_id,
                "evidence_status": evidence_status,
                "result": "pass" if passed and evidence_status == "complete" else "evidence_required",
                "policy_version": "MODEL-GATE-2026.1",
                "model_digest": model_digest,
                "dataset_digest": evaluation_digest,
                "data_nature": NATURE["11"],
            }
        )
    return pd.DataFrame(rows), [candidate_path]


VACCINE_SENSOR_ROLE_BY_CODE = {
    "LI": "freeze_preventive_internal",
    "LA": "ambient",
    "LS": "standard_carrier_internal",
}


def _phase1a_sensor_roles(zipped: ZipFile) -> dict[str, str]:
    coding_name = next(name for name in zipped.namelist() if name.endswith("Coding List Phase 1a.xlsx"))
    sheet = openpyxl.load_workbook(
        io.BytesIO(zipped.read(coding_name)), read_only=True, data_only=True
    ).active
    serial_roles: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        for serial_index, role in (
            (7, "freeze_preventive_internal"),
            (9, "ambient"),
            (11, "standard_carrier_internal"),
        ):
            serial = row[serial_index] if len(row) > serial_index else None
            if serial not in (None, ""):
                serial_roles[str(int(serial))] = role
    return serial_roles


def extract_vaccine_sensor_metadata(source_file: str, serial_roles: dict[str, str]) -> tuple[str, str]:
    """Resolve the monitor serial and role using the source coding list, then a filename-code fallback."""
    match = re.search(r"(\d{10})", source_file)
    serial = match.group(1) if match else "unavailable"
    if serial in serial_roles:
        return serial, serial_roles[serial]

    normalized = re.sub(r"[^A-Z0-9]", "", Path(source_file).stem.upper())
    for code, role in VACCINE_SENSOR_ROLE_BY_CODE.items():
        if re.search(rf"{code}(?:S?TARTED|TARTED|$)", normalized):
            return serial, role
    return serial, "unknown"


def _site_name(archive_member: str) -> str:
    value = Path(archive_member).parts[-2] if len(Path(archive_member).parts) >= 2 else "unknown"
    return re.sub(r"[_-]+(?:post)?$", "", value, flags=re.IGNORECASE).replace("_", " ").strip()


def transform_12() -> tuple[pd.DataFrame, list[Path]]:
    directory = case_dir("12")
    benchmark_path = directory / "public-benchmark.csv"
    if not benchmark_path.exists():
        current_case = directory / "case.csv"
        current = pd.read_csv(current_case)
        if {"site_code", "sensor_role", "source_file"}.issubset(current.columns):
            current.to_csv(benchmark_path, index=False, encoding="utf-8", lineterminator="\n")
        else:
            raise RuntimeError("case 12 public benchmark is missing and cannot be recovered")

    routes = [
        ("四川", "彭州市", "CN-SC-PZ-01"),
        ("四川", "都江堰市", "CN-SC-DJY-02"),
        ("云南", "安宁市", "CN-YN-AN-03"),
        ("云南", "嵩明县", "CN-YN-SM-04"),
        ("湖北", "新洲区", "CN-HB-XZ-05"),
        ("河南", "新郑市", "CN-HA-XZ-06"),
        ("陕西", "周至县", "CN-SN-ZZ-07"),
        ("浙江", "桐庐县", "CN-ZJ-TL-08"),
        ("广东", "龙门县", "CN-GD-LM-09"),
        ("福建", "闽侯县", "CN-FJ-MH-10"),
        ("辽宁", "法库县", "CN-LN-FK-11"),
        ("河北", "正定县", "CN-HE-ZD-12"),
    ]
    profile_types = ["freeze_preventive", "standard_carrier", "ambient_context"]
    start = datetime(2026, 7, 6, 7, 30, 0)
    rows = []
    for route_index, (province, county, route_id) in enumerate(routes):
        investigation_id = f"CCI-2026-{route_index + 1:03d}"
        route_status = "missing" if route_index in {2, 9} else "complete"
        calibration_status = "expired" if route_index in {4, 10} else "valid"
        handoff_status = "missing" if route_index in {0, 7} else "complete"
        for sample_index in range(30):
            temperature = 5.2 + ((sample_index % 9) - 4) * 0.28
            if 9 <= sample_index <= 13 and route_index % 3 == 0:
                temperature += 4.1
            event_time = start + timedelta(days=route_index, minutes=sample_index * 5)
            digest_source = f"{investigation_id}|{event_time.isoformat()}|{temperature:.2f}|{route_status}|{calibration_status}|{handoff_status}"
            rows.append(
                {
                    "investigation_id": investigation_id,
                    "event_id": f"CN-CC-{route_index + 1:02d}-{sample_index + 1:03d}",
                    "event_time": event_time.isoformat(),
                    "province": province,
                    "county": county,
                    "route_id": route_id,
                    "vehicle_code": f"COURSE-VEH-{route_index + 1:02d}",
                    "container_code": f"COURSE-BOX-{route_index + 1:02d}",
                    "logger_code": f"COURSE-LOGGER-{route_index + 1:02d}",
                    "temperature_c": round(temperature, 2),
                    "reference_profile_type": profile_types[route_index % len(profile_types)],
                    "route_record_status": route_status,
                    "calibration_status": calibration_status,
                    "handoff_status": handoff_status,
                    "sample_completeness": round(1 - (route_index % 4) * 0.03, 2),
                    "offline_minutes": 15 if route_index in {5, 11} and sample_index in {12, 13, 14} else 0,
                    "evidence_digest": hashlib.sha256(digest_source.encode("utf-8")).hexdigest(),
                    "policy_version": "COLD-CHAIN-INVESTIGATION-2026.1",
                    "usability_decision_allowed": False,
                    "data_nature": "deterministic-synthetic-cn-operations",
                }
            )
    return pd.DataFrame(rows), [benchmark_path]


TRANSFORMS = {
    "01": (transform_01, "一张 ¥81,768.96 的取消发票", ["取消或负数量是异常代理，不是供应商责任结论。", "人民币金额使用课程冻结汇率 1 GBP = ¥9.10 派生；不是实时汇率，也不代表中国经营事实。", "课程子集来自固定原始切片，不代表全部交易分布。"]),
    "02": (transform_02, "会员价值分层与干预实验平台", ["源文件没有绝对事件时间和金额，不能声称完整 RFM 或真实客户终身价值。", "value_segment 是课程派生的行为参与度分层。"]),
    "03": (transform_03, "本地生活客户之声研判台", ["评论样本不代表全部用户，星级分层是确定性课程抽样。", "方面标签用于证据检索和离线评测，不能直接决定需求优先级。"]),
    "04": (transform_04, "消费信贷人工复核与审计平台", ["risk_signal 只用于课程复核路由，不能自动授信或拒贷。", "历史预测样本不能证明现实公平性。"]),
    "05": (transform_05, "医院患者流程协同与异常处置中心", ["全部患者和事件均为确定性合成。", "不包含诊断、治疗建议或真实个人信息。"]),
    "06": (transform_06, "北京空气质量数据口径与证据审计台", ["缺失不是 0，相关性不是污染成因。", "case.csv 是固定抽样；完整派生数据保存在 Parquet。"]),
    "07": (transform_07, "即时零售履约领域建模与架构演进台", ["公开来源描述为 O2O grocery order set，但未证明是生产观察记录。", "它是架构与算法基准，不证明真实履约改善。"]),
    "08": (transform_08, "智慧水产多模态监测与事件响应中心", ["中国 GeoTIFF 与哥伦比亚传感器序列是独立来源。", "课程不得伪造行级关联或自动设备控制。"]),
    "09": (transform_09, "轨道交通设备 Agentic RAG 维护平台", ["已知故障窗口来自公开数据说明。", "异常与检索结果不能自动触发维修。"]),
    "10": (transform_10, "通信服务投诉升级编排台", ["case.csv 是按工信部公开类别比例生成的确定性课程运营数据，不代表真实投诉。", "FCC 公开记录单独保存在 public-reference.csv，消费者陈述未经监管机构核实。"]),
    "11": (transform_11, "企业大模型准入评审台", ["Qwen 公开元数据只证明来源锚点，课程合成评测不代表官方或企业真实准入结论。", "仓库代码许可不能替代具体模型权重和使用场景的许可核验。"]),
    "12": (transform_12, "县域冷链温度偏差调查与质量会签台", ["public-benchmark.csv 是尼泊尔公开热性能数据，case.csv 是中国县域确定性合成运营事件，两者不伪造站点映射。", "任何温度筛查或调查结案都不能自动判定产品可用、失效或放行。"]),
}

LOCAL_TRANSFORM_CASES = {"08", "09"}


def run(case_id: str) -> dict:
    started = datetime.now().astimezone().isoformat()
    if case_id in LOCAL_TRANSFORM_CASES:
        directory = case_dir(case_id)
        completed = subprocess.run(
            [sys.executable, "-B", "transform.py"],
            cwd=directory,
            check=True,
            capture_output=True,
            text=True,
        )
        source = json.loads((directory / "source.json").read_text(encoding="utf-8"))
        derived = source["derived_output"]
        receipt = {
            "case_id": case_id,
            "directory": str(directory.relative_to(ROOT)).replace("\\", "/"),
            "row_count": derived["rows"],
            "sha256": derived["sha256"],
            "data_nature": derived["data_nature"],
            "runner": "case-local-transform",
            "stdout": completed.stdout.strip(),
        }
    else:
        function, title, boundaries = TRANSFORMS[case_id]
        frame, extra = function()
        receipt = finalize(case_id, frame, title, boundaries, extra)
    receipt["started_at"] = started
    receipt["finished_at"] = datetime.now().astimezone().isoformat()
    EVIDENCE_ROOT.mkdir(parents=True, exist_ok=True)
    (EVIDENCE_ROOT / f"transform-{case_id}.json").write_text(
        json.dumps(receipt, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    return receipt


def main() -> int:
    parser = argparse.ArgumentParser(description="从固定原始数据生成课程案例数据")
    parser.add_argument("--case", choices=sorted(TRANSFORMS), required=True)
    args = parser.parse_args()
    receipt = run(args.case)
    print(json.dumps(receipt, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
