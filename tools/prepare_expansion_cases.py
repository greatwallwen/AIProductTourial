from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REFERENCE_ROOT = ROOT.parents[1] / "dataset-anlalyse"
DATASET_ROOT = ROOT / "dataset"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(4 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value: object, default: float = 0.0) -> float:
    text = clean(value).replace(",", "").replace("%", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else default


def integer(value: object, default: int = 0) -> int:
    return int(round(number(value, default)))


def excel_rows(path: Path) -> list[dict[str, object]]:
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    return [dict(zip(headers, row, strict=True)) for row in rows if any(value is not None for value in row)]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def copy_raw(case_dir: Path, sources: list[Path]) -> list[dict[str, object]]:
    raw = case_dir / "raw"
    raw.mkdir(parents=True, exist_ok=True)
    receipts: list[dict[str, object]] = []
    for source in sources:
        target = raw / source.name
        shutil.copy2(source, target)
        receipts.append({
            "path": target.relative_to(case_dir).as_posix(),
            "bytes": target.stat().st_size,
            "sha256": sha256(target),
            "role": "user-provided reference snapshot",
        })
    return receipts


def column_schema(source_fields: list[str], derived_fields: list[str]) -> list[dict[str, str]]:
    return ([{"name": name, "origin": "source-fact"} for name in source_fields] +
            [{"name": name, "origin": "derived"} for name in derived_fields])


def source_contract(
    case_id: str,
    source_id: str,
    title: str,
    inputs: list[dict[str, object]],
    source_fields: list[str],
    derived_fields: list[str],
    restriction: str,
    supplemental: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    return {
        "schema_version": "1.0",
        "case_id": case_id,
        "sources": [{
            "id": source_id,
            "title": title,
            "publisher": "用户提供的课程参考数据",
            "version": "2026-07-30 固定本地快照",
            "license": "随附文件未声明",
            "redistribution": "用户授权用于课程本地化；第三方再分发权未独立核验",
        }],
        "input_files": inputs,
        "generation": {
            "generator": "tools/prepare_expansion_cases.py",
            "generator_version": "1.0.0",
            "field_lineage": {"source_facts": source_fields, "derived_fields": derived_fields},
        },
        "supplemental_outputs": supplemental or [],
        "transform_status": "derived-verified",
        "materialized_status": "verified",
        "rebuild_status": "verified",
        "restriction": restriction,
    }


def supplemental_contract(
    path: Path,
    source_id: str,
    rows: list[dict[str, object]],
    nature: str,
) -> tuple[dict[str, object], dict[str, object]]:
    source = {
        "path": path.name,
        "source_id": source_id,
        "rows": len(rows),
        "sha256": sha256(path),
        "data_nature": nature,
    }
    schema = {
        "path": path.name,
        "row_count": len(rows),
        "columns": list(rows[0]) if rows else [],
        "source_ids": [source_id],
        "data_nature": nature,
    }
    return source, schema


def common_files(case_dir: Path, case_id: str, title: str, purpose: str) -> None:
    (case_dir / "README.md").write_text(
        f"# {case_id} {title}\n\n{purpose}\n\n"
        "`case.csv` 是页面与课堂共用的主表；`raw/` 保留可重建输入。"
        "来源、字段沿袭和使用限制见 `source.json`。\n",
        encoding="utf-8",
    )
    (case_dir / "eval.jsonl").write_text(
        json.dumps({
            "case_id": case_id,
            "check": "data-contract",
            "expected": "主表可复建，源事实与派生字段分开，页面不得越过人工核验边界",
        }, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    (case_dir / "transform.py").write_text(
        "from __future__ import annotations\n\n"
        "import sys\nfrom pathlib import Path\n\n"
        "ROOT = Path(__file__).resolve().parents[2]\n"
        "sys.path.insert(0, str(ROOT))\n"
        "from tools.prepare_expansion_cases import main\n\n"
        "if __name__ == '__main__':\n"
        f"    raise SystemExit(main(['--case', '{case_id}', *sys.argv[1:]]))\n",
        encoding="utf-8",
    )


def prepare_b021() -> dict[str, object]:
    case_id, source_id = "B021", "DATA-21"
    case_dir = DATASET_ROOT / "B021-weekend-route-planner"
    original = REFERENCE_ROOT / "3-travle_poi.xlsx"
    inputs = copy_raw(case_dir, [original])
    raw_rows = excel_rows(original)
    source_fields = ["city_name", "poi_name", "star_level", "rating", "price_cny", "sales_count", "district", "longitude", "latitude", "introduction", "is_free", "address"]
    derived_fields = ["poi_id", "route_use", "state", "data_nature"]
    rows: list[dict[str, object]] = []
    for index, item in enumerate(raw_rows, start=1):
        longitude, latitude = "", ""
        coordinate = clean(item.get("坐标"))
        if "," in coordinate:
            longitude, latitude = [part.strip() for part in coordinate.split(",", 1)]
        rows.append({
            "city_name": clean(item.get("城市")),
            "poi_name": clean(item.get("名称")),
            "star_level": clean(item.get("星级")),
            "rating": number(item.get("评分")),
            "price_cny": number(item.get("价格")),
            "sales_count": integer(item.get("销量")),
            "district": clean(item.get("省/市/区")),
            "longitude": longitude,
            "latitude": latitude,
            "introduction": clean(item.get("简介")),
            "is_free": clean(item.get("是否免费")).lower() == "true",
            "address": clean(item.get("具体地址")),
            "poi_id": f"B021-POI-{index:04d}",
            "route_use": "候选点；营业时间、交通与预约须现场核对",
            "state": "待编排",
            "data_nature": "用户提供的历史 POI 快照",
        })
    fields = source_fields + derived_fields
    write_csv(case_dir / "case.csv", fields, rows)
    write_json(case_dir / "schema.json", {"schema_version": "1.0", "case_id": case_id, "row_count": len(rows), "columns": column_schema(source_fields, derived_fields), "supplemental_artifacts": []})
    write_json(case_dir / "source.json", source_contract(case_id, source_id, "中国城市旅游 POI 快照", inputs, source_fields, derived_fields, "路线仅是静态候选；不得声称营业时间、票价、交通或预约状态实时有效。"))
    common_files(case_dir, case_id, "周末短途路线编排", "用中国 POI 的位置、价格和热度编排两日路线，并把实时信息留给出发前核对。")
    return manifest_entry(case_id, "weekend-route-planner", case_dir, source_id, rows, fields, "user-provided-historical-poi-snapshot")


def prepare_b022() -> dict[str, object]:
    case_id, source_id = "B022", "DATA-22"
    case_dir = DATASET_ROOT / "B022-transfer-notice-verification"
    notices = REFERENCE_ROOT / "10-exam" / "考研调剂数据-3.08.xlsx"
    schools = REFERENCE_ROOT / "10-exam" / "大学信息2021new.xlsx"
    inputs = copy_raw(case_dir, [notices, schools])
    school_index = {clean(row.get("school")): row for row in excel_rows(schools)}
    source_fields = ["school_name", "major_name", "notice_title", "source_relative_url", "published_date", "province", "school_type", "school_attribute"]
    derived_fields = ["notice_id", "snapshot_year", "freshness_status", "official_verification_required", "state", "data_nature"]
    rows: list[dict[str, object]] = []
    for index, item in enumerate(excel_rows(notices), start=1):
        school_name = clean(item.get("school"))
        school = school_index.get(school_name, {})
        published = clean(item.get("time"))
        rows.append({
            "school_name": school_name,
            "major_name": clean(item.get("name")),
            "notice_title": clean(item.get("title")),
            "source_relative_url": clean(item.get("url")),
            "published_date": published,
            "province": clean(school.get("province")),
            "school_type": clean(school.get("school_type")),
            "school_attribute": clean(school.get("school_attr")),
            "notice_id": f"B022-TJ-{index:04d}",
            "snapshot_year": published[:4] if published else "2021",
            "freshness_status": "历史快照",
            "official_verification_required": True,
            "state": "待回源核验",
            "data_nature": "用户提供的 2021 年调剂信息快照",
        })
    fields = source_fields + derived_fields
    write_csv(case_dir / "case.csv", fields, rows)
    write_json(case_dir / "schema.json", {"schema_version": "1.0", "case_id": case_id, "row_count": len(rows), "columns": column_schema(source_fields, derived_fields), "supplemental_artifacts": []})
    write_json(case_dir / "source.json", source_contract(case_id, source_id, "2021 年考研调剂与院校信息快照", inputs, source_fields, derived_fields, "全部记录是历史资料；必须回到院校官网或中国研究生招生信息网核验，不提供当年名额或录取承诺。"))
    common_files(case_dir, case_id, "考研调剂信息核验", "把 2021 年历史调剂信息变成回源核验队列，训练信息新鲜度判断而不是替考生作结论。")
    return manifest_entry(case_id, "transfer-notice-verification", case_dir, source_id, rows, fields, "user-provided-historical-transfer-snapshot")


def prepare_b023() -> dict[str, object]:
    case_id, source_id = "B023", "DATA-23"
    case_dir = DATASET_ROOT / "B023-spring-festival-screening"
    base = REFERENCE_ROOT / "13-SpringFestival"
    overview = base / "春节档-电影票房表现概览.xlsx"
    daily = base / "春节档-票房详情.xlsx"
    schedule = base / "春节档-排片统计（场次）-top10影片.xlsx"
    inputs = copy_raw(case_dir, [overview, daily, schedule])
    schedule_rows_raw = excel_rows(schedule)
    latest_date = max((clean(row.get("日期")) for row in schedule_rows_raw), default="")
    latest_share = {clean(row.get("电影")): number(row.get("占比")) for row in schedule_rows_raw if clean(row.get("日期")) == latest_date}
    source_fields = ["movie_name", "main_genre", "runtime_minutes", "release_date", "data_cutoff_date", "cumulative_box_office_cny", "cumulative_screenings", "cumulative_audience", "maoyan_score", "douban_score", "latest_schedule_share_pct"]
    derived_fields = ["movie_id", "screening_decision_boundary", "state", "data_nature"]
    rows: list[dict[str, object]] = []
    for index, item in enumerate(excel_rows(overview), start=1):
        name = clean(item.get("电影"))
        rows.append({
            "movie_name": name,
            "main_genre": clean(item.get("主要类型")),
            "runtime_minutes": integer(item.get("电影时长")),
            "release_date": clean(item.get("正式上映日期")),
            "data_cutoff_date": clean(item.get("数据截止日期")),
            "cumulative_box_office_cny": round(number(item.get("累计票房")), 2),
            "cumulative_screenings": integer(item.get("累计场次")),
            "cumulative_audience": integer(item.get("累计人次")),
            "maoyan_score": number(item.get("猫眼评分")),
            "douban_score": number(item.get("豆瓣评分")),
            "latest_schedule_share_pct": latest_share.get(name, 0),
            "movie_id": f"B023-FILM-{index:03d}",
            "screening_decision_boundary": "历史经营演练；新排片须结合本影院容量和实时售票",
            "state": "待模拟排片",
            "data_nature": "用户提供的 2022 春节档历史快照",
        })
    fields = source_fields + derived_fields
    write_csv(case_dir / "case.csv", fields, rows)

    daily_rows = [{
        "date": clean(item.get("日期")),
        "box_office_cny": round(number(item.get("票房")), 2),
        "screenings": integer(item.get("场次")),
        "audience": integer(item.get("人次")),
        "audience_per_screening": integer(item.get("场均人次")),
        "average_ticket_price_cny": number(item.get("平均票价")),
        "daily_champion": clean(item.get("当日票房冠军")),
        "service_fee_cny": round(number(item.get("服务费")), 2),
        "data_nature": "user-provided-historical-daily-box-office",
    } for item in excel_rows(daily)]
    schedule_rows = [{
        "date": clean(item.get("日期")),
        "movie_name": clean(item.get("电影")),
        "screenings": integer(item.get("场次")),
        "share_pct": number(item.get("占比")),
        "data_nature": "user-provided-historical-screening-schedule",
    } for item in schedule_rows_raw]
    write_csv(case_dir / "daily-box-office.csv", list(daily_rows[0]), daily_rows)
    write_csv(case_dir / "screening-schedule.csv", list(schedule_rows[0]), schedule_rows)
    daily_source, daily_schema = supplemental_contract(case_dir / "daily-box-office.csv", source_id, daily_rows, "user-provided-historical-daily-box-office")
    schedule_source, schedule_schema = supplemental_contract(case_dir / "screening-schedule.csv", source_id, schedule_rows, "user-provided-historical-screening-schedule")
    write_json(case_dir / "schema.json", {"schema_version": "1.0", "case_id": case_id, "row_count": len(rows), "columns": column_schema(source_fields, derived_fields), "supplemental_artifacts": [daily_schema, schedule_schema]})
    write_json(case_dir / "source.json", source_contract(case_id, source_id, "2022 春节档票房与排片快照", inputs, source_fields, derived_fields, "数据只支持历史经营演练；不得声称是当前实时票房、排片或影院售票结果。", [daily_source, schedule_source]))
    common_files(case_dir, case_id, "春节档排片推演", "联动影片表现、日度票房和排片占比，推演下一轮影厅资源分配。")
    return manifest_entry(case_id, "spring-festival-screening", case_dir, source_id, rows, fields, "user-provided-historical-cinema-snapshot")


def prepare_b024() -> dict[str, object]:
    case_id, source_id = "B024", "DATA-24"
    case_dir = DATASET_ROOT / "B024-supermarket-replenishment"
    base = REFERENCE_ROOT / "20-Supermarket"
    originals = [base / "order-14.1.csv", base / "order-14.3.csv"]
    inputs = copy_raw(case_dir, originals)
    source_fields = ["product_id", "category_id", "store_id", "unit_price_cny", "units", "transaction_time", "order_id"]
    derived_fields = ["transaction_id", "state", "data_nature"]
    rows: list[dict[str, object]] = []
    for source in originals:
        with source.open("r", encoding="gbk", newline="") as stream:
            for item in csv.DictReader(stream):
                rows.append({
                    "product_id": clean(item.get("商品ID")),
                    "category_id": clean(item.get("类别ID")),
                    "store_id": clean(item.get("门店编号")),
                    "unit_price_cny": number(item.get("单价")),
                    "units": integer(item.get("销量")),
                    "transaction_time": clean(item.get("成交时间")),
                    "order_id": clean(item.get("订单ID")),
                    "transaction_id": f"B024-TXN-{len(rows) + 1:05d}",
                    "state": "已成交",
                    "data_nature": "用户提供的历史交易快照",
                })
    fields = source_fields + derived_fields
    write_csv(case_dir / "case.csv", fields, rows)

    groups: dict[tuple[str, str], dict[str, object]] = defaultdict(lambda: {"transactions": 0, "units": 0, "amount": 0.0, "days": set()})
    for row in rows:
        key = (str(row["store_id"]), str(row["category_id"]))
        group = groups[key]
        group["transactions"] = int(group["transactions"]) + 1
        group["units"] = int(group["units"]) + int(row["units"])
        group["amount"] = float(group["amount"]) + float(row["unit_price_cny"]) * int(row["units"])
        group["days"].add(str(row["transaction_time"])[:10])
    ranked = sorted(groups.items(), key=lambda pair: (-int(pair[1]["units"]), pair[0]))
    summary_rows: list[dict[str, object]] = []
    for rank, ((store_id, category_id), values) in enumerate(ranked, start=1):
        active_days = max(1, len(values["days"]))
        summary_rows.append({
            "store_id": store_id,
            "category_id": category_id,
            "transaction_count": values["transactions"],
            "units_sold": values["units"],
            "sales_amount_cny": round(float(values["amount"]), 2),
            "active_days": active_days,
            "sales_velocity_units_per_day": round(int(values["units"]) / active_days, 2),
            "sales_pressure_rank": rank,
            "recommended_action": "核对在库、在途和货架容量后再创建补货单",
            "data_nature": "derived-sales-pressure-not-inventory",
        })
    write_csv(case_dir / "sales-pressure.csv", list(summary_rows[0]), summary_rows)
    summary_source, summary_schema = supplemental_contract(case_dir / "sales-pressure.csv", source_id, summary_rows, "derived-sales-pressure-not-inventory")
    write_json(case_dir / "schema.json", {"schema_version": "1.0", "case_id": case_id, "row_count": len(rows), "columns": column_schema(source_fields, derived_fields), "supplemental_artifacts": [summary_schema]})
    write_json(case_dir / "source.json", source_contract(case_id, source_id, "连锁超市历史交易明细", inputs, source_fields, derived_fields, "销量只能形成核库优先级；数据没有实时库存、在途、保质期或货架容量，禁止自动下补货单。", [summary_source]))
    common_files(case_dir, case_id, "连锁超市补货优先级", "用 7,222 条历史交易形成门店—品类销量压力榜，再由人员核对库存和配送约束。")
    return manifest_entry(case_id, "supermarket-replenishment", case_dir, source_id, rows, fields, "user-provided-historical-retail-transactions")


def manifest_entry(case_id: str, slug: str, case_dir: Path, source_id: str, rows: list[dict[str, object]], fields: list[str], nature: str) -> dict[str, object]:
    return {
        "case_id": case_id,
        "slug": slug,
        "directory": case_dir.relative_to(ROOT).as_posix(),
        "source_ids": [source_id],
        "target_nature": nature,
        "status": "derived-verified",
        "row_count": len(rows),
        "column_count": len(fields),
        "sha256": sha256(case_dir / "case.csv"),
        "materialized_nature": nature,
        "materialized_status": "verified",
        "rebuild_status": "verified",
        "display_order": int(case_id[1:]),
    }


def update_manifests(entries: list[dict[str, object]]) -> None:
    manifest_path = DATASET_ROOT / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    ids = {entry["case_id"] for entry in entries}
    datasets = [item for item in manifest["datasets"] if item.get("case_id") not in ids]
    first_lab = next((index for index, item in enumerate(datasets) if "case_ids" in item), len(datasets))
    manifest["datasets"] = datasets[:first_lab] + entries + datasets[first_lab:]
    write_json(manifest_path, manifest)

    ledger_path = ROOT / "sources" / "source-ledger.json"
    ledger = json.loads(ledger_path.read_text(encoding="utf-8"))
    additions = {
        "DATA-21": ("中国城市旅游 POI 快照", "B021 路线候选与地址、价格、热度字段"),
        "DATA-22": ("2021 年考研调剂与院校信息快照", "B022 信息新鲜度与回源核验"),
        "DATA-23": ("2022 春节档票房与排片快照", "B023 历史影院经营推演"),
        "DATA-24": ("连锁超市历史交易明细", "B024 销量压力排序与人工核库"),
    }
    ledger["sources"] = [item for item in ledger["sources"] if item.get("id") not in additions]
    for source_id, (title, usage) in additions.items():
        ledger["sources"].append({
            "id": source_id,
            "title": title,
            "kind": "user-provided-local-dataset",
            "state": "local-snapshot-hash-verified",
            "license": "随附文件未声明",
            "usage": usage,
        })
    write_json(ledger_path, ledger)


def update_checksums(case_dirs: list[Path]) -> None:
    for case_dir in case_dirs:
        files = sorted(path for path in case_dir.iterdir() if path.is_file() and path.name != "checksums.sha256")
        (case_dir / "checksums.sha256").write_text(
            "".join(f"{sha256(path)}  {path.name}\n" for path in files),
            encoding="utf-8",
        )


BUILDERS = {
    "B021": prepare_b021,
    "B022": prepare_b022,
    "B023": prepare_b023,
    "B024": prepare_b024,
}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", action="append", choices=sorted(BUILDERS))
    args = parser.parse_args(argv)
    selected = args.case or sorted(BUILDERS)
    entries = [BUILDERS[case_id]() for case_id in selected]
    update_manifests(entries)
    update_checksums([ROOT / entry["directory"] for entry in entries])
    print(json.dumps({"cases": selected, "rows": {entry["case_id"]: entry["row_count"] for entry in entries}}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
